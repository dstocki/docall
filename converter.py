import json
import argparse
from enum import Enum
from pydantic import BaseModel, Field
from typing import Optional, List, Tuple, Any

from rich.panel import Panel
from rich.console import Console
from rich.progress import track
from pathlib import Path

from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

@dataclass
class GlobalValidationState:
    invoice_numbers: List[str] = field(default_factory=list)
    issue_min_year: int = 0
    issue_max_year: int = 0
    seller_name_lc: Optional[str] = None
    sellet_tax_id: Optional[str] = None
    currency: Optional[str] = None

@dataclass
class IndividualValidationState:
    item_numbers: List[int] = field(default_factory=list)
    sum_net: float = 0.0
    sum_vat: float = 0.0
    sum_gross: float = 0.0
    vat_rate: Optional[float] = None

class Currency(str, Enum):
    EUR = "EUR"
    PLN = "PLN"
    GBP = "GBP"

class PaymentMethod(str, Enum):
    BANK_TRANSFER   = "BANK_TRANSFER"
    CASH            = "CASH"

class ServicePeriod(BaseModel):
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class InvoiceItem(BaseModel):
    number:             Optional[int]
    name_of_service:    Optional[str]
    place_of_service:   Optional[str]
    service_periods:    List[ServicePeriod] = []
    quantity:           Optional[float]
    unit_price:         Optional[float]
    net_amount:         Optional[float]
    vat_rate:           Optional[float]
    vat_amount:         Optional[float]
    gross_amount:       Optional[float]

class InvoiceBase(BaseModel):
    number:             Optional[str]
    issue_date:         Optional[str]
    service_periods:    List[ServicePeriod] = []
    place_of_service:   Optional[str]
    seller_name:        Optional[str]
    seller_address:     Optional[str]
    seller_tax_id:      Optional[str]
    buyer_name:         Optional[str]
    buyer_address:      Optional[str]
    buyer_tax_id:       Optional[str]
    currency:           Optional[Currency]
    payment_method:     Optional[PaymentMethod]
    total_net:          Optional[float]
    total_vat:          Optional[float]
    total_gross:        Optional[float]

class Invoice(InvoiceBase):
    items: List[InvoiceItem]

class ReasonedInvoiceItem(InvoiceItem):
    confidence: float       = Field(ge=0, le=1)
    warnings:   List[str]   = []

class ReasonedInvoice(InvoiceBase):
    items:              List[ReasonedInvoiceItem]
    total_confidence:   float       = Field(ge=0, le=1)
    warnings:           List[str]   = []

console = Console()

def generate_row(ws, r, idx):
    row = r + idx - 1
    thin = Side(border_style="thin", color="000000")
    ramka = Border(top=thin,left=thin,right=thin,bottom=thin)
    alignCenter = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c = ws.cell(row=row, column=1, value=idx)
    c.alignment = alignCenter
    c.border = ramka

    rich_string = CellRichText(
        f"{"Wykonane prace"}\n",
        TextBlock(InlineFont(b=True), f"BV: {"Zgorzelec ul. Sazu 3"}")
    )
    c = ws.cell(row=row, column=2, value=rich_string)
    c.alignment = alignCenter
    c.border = ramka

    c = ws.cell(row=row, column=3, value=3100000)
    c.alignment = alignCenter
    c.border = ramka
    c.number_format = '€ #,##0.00'

    c = ws.cell(row=row, column=4, value="14 Tage")
    c.alignment = alignCenter
    c.border = ramka

    dates = [
        {
            "start": "13.06.2025",
            "end": "14.06.2025"
        },
        {
            "start": "16.08.2025",
            "end": "28.08.2025"
        }
    ]

    v = ""
    for date in dates:
        if not v:
            v += f"{date.get('start')}-{date.get('end')}"
        else:
            v += f",\n{date.get('start')}-{date.get('end')}"

    c = ws.cell(row=row, column=5, value=v)
    c.alignment = alignCenter
    c.border = ramka

    c = ws.cell(row=row, column=6, value="Kenzi Spółka z o.o.\n59-900 Zgorzelec Szizu 2")
    c.alignment = alignCenter
    c.border = ramka

    c = ws.cell(row=row, column=7, value="Ja")
    c.alignment = alignCenter
    c.border = ramka

def generate_raport(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bauübersicht"

    thin = Side(border_style="thin", color="000000")
    ramka = Border(top=thin,left=thin,right=thin,bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    font_thin = Font(bold=True)
    font_def = Font(bold=False)

    widths = {'A': 6, "B": 25, 'C': 18, 'D': 15, 'E': 35, 'F': 35, 'G': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    headers = [
        ('A1', 'Nr.'),
        ('B1', 'Art der Tätigkeiten/Bauort'),
        ('D1', 'Fälligkeit der Vergütung'),
        ('E1', 'Zeitraum der Inlandstätigkeit'),
        ('E2', 'a)   Beginn'),
        ('E3', 'b)   Ende'),
        ('F1', 'Name und Anschrift des inländischen Auftraggebers'),
        ('G1', 'Ist der Auftraggeber Unternehmer?')
    ]

    for cell, text in headers:
        ws[cell] = text

    rich_string = CellRichText(
        'Auftragsumme in ',
        TextBlock(InlineFont(color='FF0000', b=True), 'EUR')
    )
    ws['C1'] = rich_string

    ws.merge_cells('A1:A3')
    ws.merge_cells('B1:B3')
    ws.merge_cells('C1:C3')
    ws.merge_cells('D1:D3')
    ws.merge_cells('F1:F3')
    ws.merge_cells('G1:G3')

    for row in range(1, 4):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = ramka
            cell.alignment = align
            if not col == 1:
                cell.font = font_thin

    cell = ws.cell(row=2, column=5)
    cell.border = ramka
    cell.alignment = align_left

    cell = ws.cell(row=3, column=5)
    cell.border = ramka
    cell.alignment = align_left

    start_row = 4
    idx = 1
    generate_row(ws, start_row, idx)
    idx += 1
    generate_row(ws, start_row, idx)
    idx += 1
    generate_row(ws, start_row, idx)
    idx += 1

    ws.cell(row=start_row+idx-1, column=2, value="SUMME:")
    summ = ws.cell(row=start_row+idx-1, column=3, value=f"=SUM(C{start_row}:C{start_row+idx-2})")
    summ.number_format = '€ #,##0.00'

    wb.save(output_path)
    print(f"File '{output_path}' saved succesfully!")

def read_context(context_dir: Path) -> Tuple[List[ReasonedInvoice], bool]:
    invoices = []

    if not context_dir.is_dir():
        print(f"Directory '{context_dir}' does not exist!")
        return [], False

    for inv_path in context_dir.glob("*.json"):
        try:
            with open(inv_path, "r", encoding="utf-8") as f:
                content = f.read()

                invoice = ReasonedInvoice.model_validate_json(content)
                invoices.append(invoice)
                print(f"Succesfully loaded: {inv_path}")

        except Exception as e:
            print(f"Parsing from JSON failed with an error: {e}")
            return [], False

    return invoices, True

def validate_context_file(
    inv: ReasonedInvoice,
    indi_data: IndividualValidationState
) -> tuple[str, bool]:
    for item in inv.items:
        if item.number is not None:
            if item.number in indi_data.item_numbers:
                return f"duplicate item number: {item.number}", False
            indi_data.item_numbers.append(item.number)

        if indi_data.vat_rate is not None:
            if item.vat_rate is not None and indi_data.vat_rate != item.vat_rate:
                return f"item VAT rate mismatch: expected {indi_data.vat_rate}, got {item.vat_rate}", False
        
        expected_item_net = item.quantity * item.unit_price if item.quantity is not None and item.unit_price is not None else None
        if expected_item_net is not None and item.net_amount is not None:
            if abs(expected_item_net - item.net_amount) > 0.01:
                return f"item net amount mismatch: expected {expected_item_net}, got {item.net_amount}", False

        expected_item_vat = item.net_amount * item.vat_rate / 100 if item.net_amount is not None and item.vat_rate is not None else None
        if expected_item_vat is not None and item.vat_amount is not None:
            if abs(expected_item_vat - item.vat_amount) > 0.01:
                return f"item VAT amount mismatch: expected {expected_item_vat}, got {item.vat_amount}", False

        expected_item_gross = item.net_amount + item.vat_amount if item.net_amount is not None and item.vat_amount is not None else None
        if expected_item_gross is not None and item.gross_amount is not None:
            if abs(expected_item_gross - item.gross_amount) > 0.01:
                return f"item gross amount mismatch: expected {expected_item_gross}, got {item.gross_amount}", False

        indi_data.sum_net += item.net_amount if item.net_amount is not None else 0
        indi_data.sum_vat += item.vat_amount if item.vat_amount is not None else 0
        indi_data.sum_gross += item.gross_amount if item.gross_amount is not None else 0

    if indi_data.sum_net != inv.total_net:
        return f"total net amount mismatch: expected {indi_data.sum_net}, got {inv.total_net}", False

    if indi_data.sum_vat != inv.total_vat:
        return f"total VAT amount mismatch: expected {indi_data.sum_vat}, got {inv.total_vat}", False

    if indi_data.sum_gross != inv.total_gross:
        return f"total gross amount mismatch: expected {indi_data.sum_gross}, got {inv.total_gross}", False

    return "", True


def validate_context(context_files: List[ReasonedInvoice]) -> tuple[str, bool]:
    global_state = GlobalValidationState()

    for inv in context_files:
        indi_state = IndividualValidationState()

        msg, valid = validate_context_file(inv, indi_state)
        if not valid:
            return msg, False
    return "", True

def main():
    parser = argparse.ArgumentParser(
        description="XLSX Generator"
    )

    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("verification.xlsx"),
        help="Path to output file"
    )

    args = parser.parse_args()

    context_dir = Path("invoices_context")
    output_dir = args.xlsx

    console.print(
        Panel(
f"""[bold]XLSX Generator[/bold]

[cyan]Parameters:[/cyan]
• Context directory:    {context_dir}
• Output file:          {output_dir}""",
            title="Setup",
        )
    )

    out_suffix = output_dir.suffix.lower()
    if out_suffix != ".xlsx":
        print(f"Incorrect file type {out_suffix}!")
        return

    if not context_dir.exists():
        print("Context directory does not exist!")
        return

    ctx_files = list(context_dir.iterdir())

    if len(ctx_files) == 0:
        print("Context directory is empty!")
        return

    context_files, ok = read_context(context_dir)
    if not ok:
        print("Generation stopped.")
        return

    print(f"Reading successful. Context files: {len(context_files)}")
    
    msg, valid = validate_context(context_files)
    if not valid:
        print(f"Validation failed: {msg}")
        return

    print("Validation successful. Generating XLSX report...")

    # data = transform_context(context_files)
    generate_raport(output_dir)

if __name__ == "__main__":
    main()