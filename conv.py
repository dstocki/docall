from collections import defaultdict
from http import client
import json
import argparse
from enum import Enum
from main import Address, Currency, PaymentMethod, InvoiceItem, InvoiceBase, Invoice, ReasonedInvoice, ReasonedInvoiceItem
from pydantic import BaseModel, Field
from typing import Dict, Optional, List, Tuple, Any, Set
from datetime import datetime, timedelta

from openai import OpenAI
from rich.panel import Panel
from rich.console import Console
from rich.progress import track
from pathlib import Path

from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

client = OpenAI()

SYSTEM_PROMPT = Path("prompts/service_translation.txt").read_text(
    encoding="utf-8"
)

@dataclass
class GlobalValidationState:
    invoice_numbers: List[str] = field(default_factory=list)
    issue_min_year: int = 0
    issue_max_year: int = 0
    seller_name: Optional[str] = None

@dataclass
class IndividualValidationState:
    item_numbers: List[int] = field(default_factory=list)
    sum_net: float = 0.0
    sum_vat: float = 0.0
    sum_gross: float = 0.0
    vat_is_set: bool = False
    vat_amount: Optional[float] = None

console = Console()

def generate_row(ws, r, idx, data):
    row = r + idx - 1
    thin = Side(border_style="thin", color="000000")
    ramka = Border(top=thin,left=thin,right=thin,bottom=thin)
    alignCenter = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    c = ws.cell(row=row, column=1, value=idx)
    c.alignment = alignCenter
    c.border = ramka

    rich_string = CellRichText(
        f"{data.get('service_name', 'N/A')}\n",
        TextBlock(InlineFont(b=True), f"BV: {data.get('bv', 'N/A')}")
    )
    c = ws.cell(row=row, column=2, value=rich_string)
    c.alignment = alignCenter
    c.border = ramka

    c = ws.cell(row=row, column=3, value=data.get('sum', 0))
    c.alignment = alignCenter
    c.border = ramka
    c.number_format = '€ #,##0.00'

    c = ws.cell(row=row, column=4, value=data.get('payment_term', 'N/A'))
    c.alignment = alignCenter
    c.border = ramka

    # Okresy to teraz jeden gotowy string stworzony przez merge_date_intervals
    c = ws.cell(row=row, column=5, value=data.get('periods', ''))
    c.alignment = alignCenter
    c.border = ramka

    c = ws.cell(row=row, column=6, value=data.get('b_name', 'N/A'))
    c.alignment = alignCenter
    c.border = ramka

    c = ws.cell(row=row, column=7, value=data.get('is_entrepreneur', 'N/A'))
    c.alignment = alignCenter
    c.border = ramka

def generate_report(data: List[Dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bauübersicht"

    thin = Side(border_style="thin", color="000000")
    ramka = Border(top=thin,left=thin,right=thin,bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    font_thin = Font(bold=True)
    font_def = Font(bold=False)

    widths = {'A': 6, "B": 25, 'C': 18, 'D': 15, 'E': 35, 'F': 35, 'G': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    headers = [
        ('A1', 'Nr.'),
        ('B1', 'Art der Tätigkeiten/Bauort'),
        ('D1', 'Fälligkeit der Vergütung'),
        ('E1', 'Zeitraum der Inlandstätigkeit'),
        ('E2', 'a)   Beginn'),
        ('E3', 'b)   Ende'),
        ('F1', 'Name und Anschrift des inländischen Auftraggebers'),
        ('G1', 'Ist der Auftraggeber Unternehmer?')
    ]

    for cell, text in headers:
        ws[cell] = text

    rich_string = CellRichText(
        'Auftragsumme in ',
        TextBlock(InlineFont(color='FF0000', b=True), 'EUR')
    )
    ws['C1'] = rich_string

    ws.merge_cells('A1:A3')
    ws.merge_cells('B1:B3')
    ws.merge_cells('C1:C3')
    ws.merge_cells('D1:D3')
    ws.merge_cells('F1:F3')
    ws.merge_cells('G1:G3')

    for row in range(1, 4):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = ramka
            cell.alignment = align
            if not col == 1:
                cell.font = font_thin

    cell = ws.cell(row=2, column=5)
    cell.border = ramka
    cell.alignment = align_left

    cell = ws.cell(row=3, column=5)
    cell.border = ramka
    cell.alignment = align_left

    start_row = 4
    idx = 1
    for d in data:
        generate_row(ws, start_row, idx, d)
        idx += 1

    ws.cell(row=start_row+idx-1, column=2, value="SUMME:")
    summ = ws.cell(row=start_row+idx-1, column=3, value=f"=SUM(C{start_row}:C{start_row+idx-2})")
    summ.number_format = '€ #,##0.00'

    wb.save(output_path)
    print(f"File '{output_path}' saved succesfully!")

def read_context(context_dir: Path) -> Tuple[List[ReasonedInvoice], bool]:
    invoices = []
    file_mapping = {}

    if not context_dir.is_dir():
        print(f"Directory '{context_dir}' does not exist!")
        return [], {}, False

    for inv_path in context_dir.glob("*.json"):
        try:
            with open(inv_path, "r", encoding="utf-8") as f:
                content = f.read()

            invoice = ReasonedInvoice.model_validate_json(content)
            invoices.append(invoice)

            file_mapping[id(invoice)] = inv_path
            print(f"Succesfully loaded: {inv_path}")

        except Exception as e:
            print(f"Parsing from JSON failed with an error: {e}")
            return [], {}, False

    return invoices, file_mapping, True

def is_correct_full_date(date: str) -> bool:
    try:
        _ = datetime.strptime(date, "%d.%m.%Y")
        return True
    except ValueError:
        return False

def split_date(date: str) -> tuple[bool, list[int]]:
    try:
        dt = datetime.strptime(date, "%d.%m.%Y")
        return True, [dt.year, dt.month, dt.day]
    except ValueError:
        pass

    try:
        dt = datetime.strptime(date, "%m.%Y")
        return True, [dt.year, dt.month]
    except ValueError:
        return False, []
    
def split_address(addr: str) -> tuple[bool, dict]:
    try:
        parts = [p.strip() for p in addr.split(',')]
        if len(parts) != 3:
            return False, {}
        
        street_part, postal_part, country = parts
        street, number = street_part.rsplit(' ', 1)
        code, city = postal_part.split(' ', 1)

        return True, {
            "street": street,
            "number": number,
            "code": code,
            "city": city,
            "country": country
        }
    except ValueError:
        return False, {}

def validate_context_file(
    inv: ReasonedInvoice,
    indi_data: IndividualValidationState
) -> tuple[str, bool]:
    err_msg = ""
    err_item = None

    for item in inv.items:
        err_item = item.number

        # --- INVOICE ITEM NUMBER ---
        if item.number is None:
            err_msg = "item number missing"
            break
        if item.number in indi_data.item_numbers:
            err_msg = f"duplicate item number: {item.number}"
            break
        indi_data.item_numbers.append(item.number)

        # --- INVOICE ITEM NAME OF SERVICE ---
        if item.name_of_service is None:
            err_msg = "item name of service missing"
            break
        if len(item.name_of_service.strip()) == 0:
            err_msg = "item name of service is empty"
            break

        # --- INVOICE ITEM PLACE OF SERVICE ---
        if item.place_of_service is None:
            err_msg = "item place of service missing"
            break
        if item.place_of_service.country is None:
            err_msg = "item place of service country missing"
            break
        if item.place_of_service.country != "Deutschland":
            err_msg = f"item place of service country is not Deutschland: {item.place_of_service.country}"
            break
        if item.place_of_service.city is None:
            err_msg = "item place of service city missing"
            break
        if not item.place_of_service.street and item.place_of_service.building_number:
            err_msg = "item place of service street missing"
            break
        if item.place_of_service.street and not item.place_of_service.building_number:
            err_msg = "item place of service building number missing"
            break

        # --- INVOICE ITEM SERVICE PERIODS ---
        if not item.service_periods:
            err_msg = "item service periods object missing"
            break
        if len(item.service_periods) == 0:
            err_msg = "item service periods list is empty"
            break

        err_serv_perd_msg = ""
        err_serv_perd = None
        for i, serv_perds in enumerate(item.service_periods):
            err_serv_perd = i
            if serv_perds.start_date is None and serv_perds.end_date is None:
                err_serv_perd_msg = f"missing start and end date"
                break
            if serv_perds.start_date is None or len(serv_perds.start_date) == 0:
                err_serv_perd_msg = f"missing start date"
                break
            if serv_perds.end_date is None or len(serv_perds.end_date) == 0:
                err_serv_perd_msg = f"missing end date"
                break
            ok, start_split = split_date(serv_perds.start_date)
            if not ok:
                err_serv_perd_msg = f"incorrect start date: {serv_perds.start_date}"
                break
            ok, end_split = split_date(serv_perds.end_date)
            if not ok:
                err_serv_perd_msg = f"incorrect end date: {serv_perds.end_date}"
                break
            if len(start_split) != len(end_split):
                err_serv_perd_msg = f"start and end date in different formats: {serv_perds.start_date}-{serv_perds.end_date}"
                break
            if start_split > end_split:
                err_serv_perd_msg = f"start date {serv_perds.start_date} is after end date {serv_perds.end_date}"
                break
        if err_serv_perd_msg:
            err_msg = f"service period {err_serv_perd}: {err_serv_perd_msg}"
            break

        # --- INVOICE ITEM TOTALS ---
        if item.net_amount is None:
            err_msg = "net amount missing"
            break
        if item.net_amount <= 0.0:
            err_msg = "net amount must be greater than 0.00"
            break
        if item.vat_amount is None:
            err_msg = "VAT amount missing"
            break
        if item.vat_amount < 0.0:
            err_msg = "VAT amount must be greater than or equal to 0.00"
            break
        if item.gross_amount is None:
            err_msg = "gross amount missing"
            break
        if item.gross_amount <= 0.0:
            err_msg = "gross amount must be greater than 0.00"
            break
        if not indi_data.vat_is_set:
            indi_data.vat_amount = item.vat_amount
            indi_data.vat_is_set = True
        if indi_data.vat_amount != 0.0 and item.vat_amount == 0.0:
            err_msg = f"VAT rate mismatch: expected VAT, got no VAT"
            break
        if indi_data.vat_amount == 0.0 and item.vat_amount != 0.0:
            err_msg = f"VAT rate mismatch: expected no VAT, got VAT"
            break
        expected_item_gross = round(item.net_amount + item.vat_amount, 2)
        if abs(expected_item_gross - item.gross_amount) > 0.01:
            err_msg = f"gross amount mismatch: expected {expected_item_gross}, got {item.gross_amount}"
            break

        indi_data.sum_net += item.net_amount
        indi_data.sum_vat += item.vat_amount
        indi_data.sum_gross += item.gross_amount

    if err_msg:
        return f"item nr {err_item}: {err_msg}", False

    indi_data.sum_net = round(indi_data.sum_net, 2)
    indi_data.sum_vat = round(indi_data.sum_vat, 2)
    indi_data.sum_gross = round(indi_data.sum_gross, 2)

    if abs(indi_data.sum_net - inv.total_net) > 0.01:
        return f"total net amount mismatch: expected {indi_data.sum_net}, got {inv.total_net}", False
    if abs(indi_data.sum_vat - inv.total_vat) > 0.01:
        return f"total VAT amount mismatch: expected {indi_data.sum_vat}, got {inv.total_vat}", False
    if abs(indi_data.sum_gross - inv.total_gross) > 0.01:
        return f"total gross amount mismatch: expected {indi_data.sum_gross}, got {inv.total_gross}", False

    return "", True

def format_address(addr) -> str:
    if addr is None:
        return ""

    street_part = f"{addr.street or ''} {addr.building_number or ''}".strip()
    city_part = f"{addr.postal_code or ''} {addr.city or ''}".strip()
    country_part = addr.country or ""
    
    parts = [p for p in [street_part, city_part, country_part] if p]
    return ", ".join(parts)

class ServiceSummary(BaseModel):
    summary: str

def summarize_services_with_ai(services: Set[str]) -> str:
    if not services:
        return ""

    services_text = ", ".join(services)
    
    try:
        response = client.responses.parse(
            model="gpt-5.4",
            input=[
                {
                    "role": "system",
                    "content": SYSTEM_PROMPT,
                },
                {
                    "role": "user",
                    "content": f"Services to summarize: {services_text}",
                },
            ],
            text_format=ServiceSummary,
        )
        
        return response.output_parsed.summary
    except Exception as e:
        print(f"AI Summarization failed: {e}")
        return "Verschiedene Dienstleistungen"

def merge_date_intervals(periods: List[dict]) -> str:
    if not periods:
        return ""
    
    intervals = []
    unparsed = set()
    
    for period in periods:
        start_str = period.get("start")
        end_str = period.get("end")
        
        if not start_str:
            continue
            
        try:
            start_dt = datetime.strptime(start_str, "%d.%m.%Y")
            end_dt = datetime.strptime(end_str, "%d.%m.%Y") if end_str else start_dt
            
            if end_dt < start_dt:
                start_dt, end_dt = end_dt, start_dt
                
            intervals.append([start_dt, end_dt])
        except ValueError:
            s = f"{start_str}-{end_str}" if end_str and start_str != end_str else start_str
            unparsed.add(s)
            
    # Chronologiczne sortowanie po poprawnym zamienieniu na daty
    intervals.sort(key=lambda x: x[0])
    
    merged = []
    for current in intervals:
        if not merged:
            merged.append(current)
        else:
            last_start, last_end = merged[-1]
            current_start, current_end = current
            
            # Łączymy jeśli nakładają się na siebie lub następują dzień po dniu
            if current_start <= last_end + timedelta(days=1):
                merged[-1][1] = max(last_end, current_end)
            else:
                merged.append(current)
                
    result_strings = []
    for start, end in merged:
        if start == end:
            result_strings.append(start.strftime("%d.%m.%Y"))
        else:
            result_strings.append(f"{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}")
            
    result_strings.extend(list(unparsed))
    return ",\n".join(result_strings)

def transform_context(context_files: List[ReasonedInvoice]) -> List[dict]:
    grouped_data = defaultdict(lambda: {
        "sum": 0.0,
        "services": set(),
        "payment_methods": set(),
        "periods": [],
        "buyer_full_address": "",
        "is_entrepreneur": "Nein"
    })

    for inv in context_files:
        buyer_name = inv.buyer_name or ""
        buyer_addr_str = format_address(inv.buyer_address)
        buyer_full = f"{buyer_name}\n{buyer_addr_str}".strip()

        is_entrepreneur = "Ja" if inv.total_vat == 0.0 else "Nein"

        for item in inv.items:
            bv_str = format_address(item.place_of_service)
            group_key = (buyer_name, bv_str)

            grouped_data[group_key]["sum"] += (item.gross_amount or 0.0)

            if item.name_of_service:
                grouped_data[group_key]["services"].add(item.name_of_service)
            if inv.payment_method:
                grouped_data[group_key]["payment_methods"].add(inv.payment_method.value)
            if item.service_periods:
                for period in item.service_periods:
                    grouped_data[group_key]["periods"].append({
                        "start": period.start_date,
                        "end": period.end_date
                    })

            grouped_data[group_key]["buyer_full_address"] = buyer_full
            grouped_data[group_key]["is_entrepreneur"] = is_entrepreneur

    trans_data = []

    for (b_name, bv_str), data in grouped_data.items():
        payment_methods_str = ", ".join(data["payment_methods"])
        if PaymentMethod.BANK_TRANSFER.value in payment_methods_str:
            payment_term = "14 Tage"
        elif PaymentMethod.CASH.value in payment_methods_str:
            payment_term = "Bar zahlen"
        else:
            payment_term = "14 Tage"

        ai_summary = summarize_services_with_ai(data["services"])
        formatted_periods = merge_date_intervals(data["periods"])

        row = {
            "service_name": ai_summary,
            "bv": bv_str,
            "sum": round(data["sum"], 2),
            "payment_term": payment_term,
            "periods": formatted_periods,
            "b_name": data["buyer_full_address"],
            "is_entrepreneur": data["is_entrepreneur"]
        }

        trans_data.append(row)

    return trans_data

def validate_context(context_files: List[ReasonedInvoice], file_mapping: dict[int, str]) -> tuple[str, bool]:
    global_state = GlobalValidationState()

    err_msg = ""
    err_inv = None

    for inv in context_files:
        err_inv = inv

        indi_state = IndividualValidationState()

        # --- INVOICE NUMBER ---
        if not inv.number:
            err_msg = "invoice number missing"
            break

        # --- INVOICE ISSUE DATE ---
        if not inv.issue_date:
            err_msg = "issue date missing"
            break
        if not is_correct_full_date(inv.issue_date):
            err_msg = f"incorrect format or values of issue date {inv.issue_date}"
            break

        # --- INVOICE CURRENCY ---
        if not inv.currency:
            err_msg = "currency missing"
            break
        if inv.currency not in Currency.__members__.values():
            err_msg = f"unsupported currency: {inv.currency}"
            break

        # --- INVOICE PAYMENT METHOD ---
        if not inv.payment_method:
            err_msg = "payment method missing"
            break
        if inv.payment_method not in PaymentMethod.__members__.values():
            err_msg = f"unsupported payment method: {inv.payment_method}"
            break

        # --- INVOICE TOTALS ---
        if inv.total_net is None:
            err_msg = "total net amount missing"
            break
        if inv.total_vat is None:
            err_msg = "total VAT amount missing"
            break
        if inv.total_gross is None:
            err_msg = "total gross amount missing"
            break
        if inv.total_net <= 0.0:
            err_msg = "total net amount must be greater than 0.00"
            break
        if inv.total_vat < 0.0:
            err_msg = "total VAT amount must be greater than or equal to 0.00"
            break
        if inv.total_gross <= 0.0:
            err_msg = "total gross amount must be greater than 0.00"
            break

        # --- INVOICE SELLER DETAILS ---
        if not inv.seller_name:
            err_msg = "seller name missing"
            break
        if not inv.seller_address:
            err_msg = "seller address object missing"
            break
        if not inv.seller_address.country:
            err_msg = "seller address country missing"
            break
        if not inv.seller_address.city:
            err_msg = "seller address city missing"
            break
        if not inv.seller_address.postal_code:
            err_msg = "seller address postal code missing"
            break
        if not inv.seller_address.street:
            err_msg = "seller address street missing"
            break
        if not inv.seller_address.building_number:
            err_msg = "seller address building number missing"
            break

        # --- INVOICE BUYER DETAILS ---
        if not inv.buyer_name:
            err_msg = "buyer name missing"
            break
        if not inv.buyer_address:
            err_msg = "buyer address object missing"
            break
        if not inv.buyer_address.country:
            err_msg = "buyer address country missing"
            break
        if not inv.buyer_address.city:
            err_msg = "buyer address city missing"
            break
        if not inv.buyer_address.postal_code:
            err_msg = "buyer address postal code missing"
            break
        if not inv.buyer_address.street:
            err_msg = "buyer address street missing"
            break
        if not inv.buyer_address.building_number:
            err_msg = "buyer address building number missing"
            break

        # --- INVOICE ITEMS ---
        if not inv.items:
            err_msg = "invoice items object missing"
            break
        if len(inv.items) == 0:
            err_msg = "invoice items list is empty"
            break

        # --- GLOBAL INVOICE VALIDATION ---
        if inv.number in global_state.invoice_numbers:
            err_msg = f"duplicated invoice number: {inv.number}"
            break
        global_state.invoice_numbers.append(inv.number)

        _, parts = split_date(inv.issue_date)
        curr_year = parts[0]

        if global_state.issue_max_year == 0:
            global_state.issue_min_year = curr_year
            global_state.issue_max_year = curr_year
        else:
            global_state.issue_min_year = min(global_state.issue_min_year, curr_year)
            global_state.issue_max_year = max(global_state.issue_max_year, curr_year)

        if abs(global_state.issue_max_year - global_state.issue_min_year) > 1:
            err_msg = f"invoices spread on more than 2 different years"
            break
        
        if not global_state.seller_name:
            global_state.seller_name = inv.seller_name
        
        if global_state.seller_name != inv.seller_name:
            err_msg = f"seller name mismatch: expected {global_state.seller_name}, got {inv.seller_name}"
            break

        msg, valid = validate_context_file(inv, indi_state)
        if not valid:
            err_msg = f"{msg}"
            break

    if err_msg:
        return f"invoice {file_mapping.get(id(err_inv), 'N/A')}: {err_msg}", False
    return "", True

def main():
    parser = argparse.ArgumentParser(
        description="XLSX Generator"
    )

    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("verification.xlsx"),
        help="Path to output file"
    )

    args = parser.parse_args()

    context_dir = Path("invoices_context")
    output_dir = args.xlsx

    console.print(
        Panel(
f"""[bold]XLSX Generator[/bold]

[cyan]Parameters:[/cyan]
• Context directory:    {context_dir}
• Output file:          {output_dir}""",
            title="Setup",
        )
    )

    out_suffix = output_dir.suffix.lower()
    if out_suffix != ".xlsx":
        print(f"Incorrect file type {out_suffix}!")
        return

    if not context_dir.exists():
        print("Context directory does not exist!")
        return

    ctx_files = list(context_dir.iterdir())

    if len(ctx_files) == 0:
        print("Context directory is empty!")
        return

    context_files, file_mapping, ok = read_context(context_dir)
    if not ok:
        print("Generation stopped.")
        return

    print(f"Reading successful. Context files: {len(context_files)}")
    
    msg, valid = validate_context(context_files, file_mapping)
    if not valid:
        print(f"Validation failed: {msg}")
        return

    print("Validation successful. Generating XLSX report...")

    data = transform_context(context_files)
    generate_report(data, output_dir)

if __name__ == "__main__":
    main()